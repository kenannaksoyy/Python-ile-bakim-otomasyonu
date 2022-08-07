import datetime
from cgitb import text
from tkinter import font

from openpyxl import Workbook,load_workbook
import datetime as dt
from tkinter import *
import tkinter.messagebox as tm
import xlsxwriter



class Usta:
    def __init__(self, id, isim, soyisim, alan, dTarih, yaş, bTarih, işYıl, skor):
        self.id, self.isim, self.soyisim, self.alan = str(id), isim, soyisim, alan
        self.dTarih, self.yaş, self.bTarih, self.işYıl, self.skor = dTarih, yaş, bTarih, işYıl, int(skor)
class Kullanıcı:
    def __init__(self, kullanıcıİsim, şifre):
        self.kullanıcıİsim, self.şifre = kullanıcıİsim, şifre
class Arıza:
    def __init__(self,arızaKodu, arızaAlan, arızaSüre, arızaİsim):
        self.arızaKodu, self.arızaAlan, self.arızaSüre, self.arızaİsim = arızaKodu, arızaAlan, int(arızaSüre), arızaİsim
        if arızaAlan == "Hidrolik":
            self.arızaSkor = 8 * self.arızaSüre
        if arızaAlan == "Elektrik":
            self.arızaSkor = 9 * self.arızaSüre
        if arızaAlan == "Kalıp":
            self.arızaSkor = 10 * self.arızaSüre
        if arızaAlan == "Mekanik":
            self.arızaSkor = 8 * self.arızaSüre

def excel_veri_çekme(dosyaİsim, dosyaSayfaİsim):
    wb = load_workbook(dosyaİsim)
    ws = wb[dosyaSayfaİsim]
    satirSayısı = ws.max_row - 1
    sütunSayısı = ws.max_column
    bilgiMatrix = [["A" for x in range(sütunSayısı)] for y in range(satirSayısı)]
    satirSayac = 0
    sütunSayac = 0
    for satır in range(2, ws.max_row + 1):
        sütunSayac = 0
        for sütun in range(1, ws.max_column + 1):
            bilgiMatrix[satirSayac][sütunSayac] = str(ws.cell(satır, sütun).value)
            sütunSayac = sütunSayac + 1
        satirSayac = satirSayac + 1
    wb.close()
    return bilgiMatrix

def zaman_hesaplama(tarih):
    tarih = dt.datetime.strptime(tarih,"%d.%m.%Y")
    bugün = dt.datetime.now()
    fark = bugün - tarih
    return  fark.days//365

def id_isim_arama_kontrol(ustalarList, id, isim, soyisim):
    for n in ustalarList:
        if id == n.id and isim == n.isim and soyisim == n.soyisim:
            return 1
        else:
            return 0

def arıza_skorlama(arızalarList, arızaKodu):
    for n in arızalarList:
        if arızaKodu == n.arızaKodu:
            return n.arızaSkor
        else:
            return 0

def mola_skorlama(molaSüresi, ustalarList, id, molaTür):
    for n in ustalarList:
        if id == n.id:
            if molaTür == "Öğle":
                skor = (30 - molaSüresi) * 3
                return skor
            if molaTür == "Sabah" and molaTür == "Akşam":
                skor = (15 - molaSüresi) * 3
                return skor

class Arayüz:
    def __init__(self, pencere):
        self.pencere = pencere
        self.pencere.title("Legrand Bakım Otomosyonu")
        self.pencere.geometry("600x600")

        icon = PhotoImage(file = "LogoIcon.png")
        self.pencere.iconphoto(False,icon)

        self.resim = PhotoImage(file = "LegrandLogin.png")
        Label(self.pencere, image = self.resim).place(x=60,y=240)

        self.resim1 = PhotoImage(file="Turkiye.png")
        Label(self.pencere, image=self.resim1).place(x=0, y=0)

        self.resim2 = PhotoImage(file="Fransa.png")
        Label(self.pencere, image=self.resim2).place(x=425, y=0)

        Label(self.pencere,text="Kullanıcı Adı").place(x = 195, y = 80)
        Label(self.pencere,text="Parola").place(x = 195, y = 110)

        self.kullanıcı = Entry(self.pencere)
        self.kullanıcı.place(x = 280, y = 80)
        self.parola = Entry(self.pencere, show = "*")
        self.parola.place(x = 280, y = 110)

        self.girişButon = Button(self.pencere,text="Giriş",width=5,background="blue",
        command = self.giriş_kontrol).place(x = 360, y=140)
        self.çıkışButon = Button(self.pencere,text ="Çıkış",width=5,background="red",
        command = self.pencere.quit).place(x=280, y=140)

    def giriş_kontrol(self):
        kullanıcıBilgi =  self.kullanıcı.get()
        parolaBilgi = self.parola.get()

        if not self.kullanıcı.get() or not self.parola.get():
            bosluk = tm.askyesno(title="Hata Mesajı",message="Gerekli Yerler Yazılmadı Lütfen Tekrar Giriş Yapın")
            if bosluk == False:
                self.pencere.destroy()

        for n in kullanıcılarList:
            if n.kullanıcıİsim == kullanıcıBilgi and n.şifre == parolaBilgi:
                self.doğru = tm.showinfo("Giriş Mesajı", "Giriş Başarılı Hoşgeldiniz")
                self.anasayfaGidiş()

        for n in kullanıcılarList:
            if n.kullanıcıİsim != kullanıcıBilgi and n.şifre == parolaBilgi:
                tm.showerror("Hata Mesajı", "Kullanıcı Adını Yanlış Girdiniz")
        for n in kullanıcılarList:
            if n.kullanıcıİsim == kullanıcıBilgi and n.şifre != parolaBilgi:
                tm.showerror("Hata Mesajı", "Parolayı Yanlış Girdiniz")



    def anasayfaGidiş(self):
        for etiket in self.pencere.winfo_children():
            etiket.destroy()
        self.anasayfa()

    def yokedici(self):
        for etiket in self.pencere.winfo_children():
            etiket.destroy()

    def usta_bilgileri_sayfa(self,id):
        self.yokedici()
        self.resim1 = PhotoImage(file="Turkiye.png")
        Label(self.pencere, image=self.resim1).place(x=0, y=0)

        self.resim2 = PhotoImage(file="Fransa.png")
        Label(self.pencere, image=self.resim2).place(x=425, y=0)

        self.resim3 = PhotoImage(file="UstaResim.png")
        Label(self.pencere, image=self.resim3).place(x=0, y=150)

        for n in ustalarList:
            if n.id == id:
                Label(self.pencere, text="İD :", font = "Verdana 12").place(x=175, y=150)
                Label(self.pencere, text=n.id, font = "Verdana 12").place(x=310, y=150)
                Label(self.pencere, text="İsim :", font="Verdana 12").place(x=175, y=175)
                Label(self.pencere, text=n.isim, font="Verdana 12").place(x=310, y=175)
                Label(self.pencere, text="Soyisim :", font="Verdana 12").place(x=175, y=200)
                Label(self.pencere, text=n.soyisim, font="Verdana 12").place(x=310, y=200)
                Label(self.pencere, text="Alan :", font="Verdana 12").place(x=175, y=225)
                Label(self.pencere, text=n.alan, font="Verdana 12").place(x=310, y=225)
                Label(self.pencere, text="Doğum Tarih :", font="Verdana 12").place(x=175, y=250)
                Label(self.pencere, text=n.dTarih, font="Verdana 12").place(x=310, y=250)
                Label(self.pencere, text="Yaş :", font="Verdana 12").place(x=175, y=275)
                Label(self.pencere, text=n.yaş, font="Verdana 12").place(x=310, y=275)
                Label(self.pencere, text="Başlama Tarih :", font="Verdana 12").place(x=175, y=300)
                Label(self.pencere, text=n.bTarih, font="Verdana 12").place(x=310, y=300)
                Label(self.pencere, text="İş Yılı :", font="Verdana 12").place(x=175, y=325)
                Label(self.pencere, text=n.işYıl, font="Verdana 12").place(x=310, y=325)
                Label(self.pencere, text="Skor :", font="Verdana 12").place(x=175, y=350)
                Label(self.pencere, text=n.skor, font="Verdana 12").place(x=310, y=350)

        self.UBGAnasayfaButon=Button(self.pencere, text="Anasayfa", width=8,
        command=self.anasayfaGidiş, background="blue").place(x=480, y=400)
        self.UBGÇıkışButon = Button(self.pencere, text="Çıkış", width=5, background="red",
        command=self.pencere.quit).place(x=15, y=400)

    def usta_bilgileri_sayfa_kontrol(self):
        kontrolİdGirdi = self.idGirdi.get()
        kontrolİsimGirdi = self.isimGirdi.get()
        kontrolSoyisimGirdi = self.soyisimGirdi.get()

        if not self.idGirdi.get() or not self.isimGirdi.get() or not self.soyisimGirdi.get():
            bosluk = tm.showerror("Hata Mesajı","Gerekli Yerler Yazılmadı Lütfen Tekrar Giriş Yapın")
            if bosluk == False:
                self.pencere.destroy()
        for n in ustalarList:
            if n.id != kontrolİdGirdi and n.isim == kontrolİsimGirdi and n.soyisim == kontrolSoyisimGirdi:
                tm.showerror("Kontrol Mesajı", "İD Yanlış Girdiniz")
        for n in ustalarList:
            if n.id == kontrolİdGirdi and n.isim != kontrolİsimGirdi and n.soyisim == kontrolSoyisimGirdi:
                tm.showerror("Kotrol Mesajı", "İsim Yanlış Girdiniz")
        for n in ustalarList:
            if n.id == kontrolİdGirdi and n.isim == kontrolİsimGirdi and n.soyisim != kontrolSoyisimGirdi:
                tm.showerror("Kontrol Mesajı", "Soyisim Yanlış Girdiniz")
        for n in ustalarList:
            if n.id == kontrolİdGirdi and n.isim == kontrolİsimGirdi and n.soyisim == kontrolSoyisimGirdi:
                self.ustaKontrolDoğru = tm.showinfo("Kontrol Mesajı", "Usta Kayıdı Bulundu")
                self.usta_bilgileri_sayfa(kontrolİdGirdi)

    def arıza_bilgileri_sayfa_kontrol(self):
        kontrolArızaKodGirdi=self.arızaKodGirdi.get()
        sayac = 0

        if not self.arızaKodGirdi.get():
            bosluk=tm.showerror("Hata Mesajı", "Arıza Kodunu Girilmedi Lütfen Tekrar Giriş Yapın")
            if bosluk == False:
                self.pencere.destroy()
        for n in arızalarList:
            if n.arızaKodu == kontrolArızaKodGirdi:
                sayac = 1
                self.arızaKontrolDogru = tm.showinfo("Kontrol Mesajı", "Arıza Kayıdı Bulundu")
                self.arıza_bilgileri_sayfa(kontrolArızaKodGirdi)
        if sayac == 0:
            yanlıs=tm.showerror("Kontrol Mesajı", "Arıza Kaydı Bulunmadı")


    def arıza_bilgileri_sayfa(self,arızaKod):
        self.yokedici()
        self.resim1 = PhotoImage(file="Turkiye.png")
        Label(self.pencere, image=self.resim1).place(x=0, y=0)

        self.resim2 = PhotoImage(file="Fransa.png")
        Label(self.pencere, image=self.resim2).place(x=425, y=0)

        self.resim3 = PhotoImage(file="Arıza.png")
        Label(self.pencere, image=self.resim3).place(x=0, y=160)

        for n in arızalarList:
            if n.arızaKodu == arızaKod:
                Label(self.pencere, text="Arıza Kodu :", font="Verdana 12").place(x=175, y=225)
                Label(self.pencere, text=n.arızaKodu, font="Verdana 12").place(x=310, y=225)
                Label(self.pencere, text="Arıza Alan :", font="Verdana 12").place(x=175, y=250)
                Label(self.pencere, text=n.arızaAlan, font="Verdana 12").place(x=310, y=250)
                Label(self.pencere, text="Arıza Süre :", font="Verdana 12").place(x=175, y=275)
                Label(self.pencere, text=n.arızaSüre, font="Verdana 12").place(x=310, y=275)
                Label(self.pencere, text="Arıza :", font="Verdana 12").place(x=175, y=300)
                Label(self.pencere, text=n.arızaİsim, font="Verdana 12").place(x=310, y=300)
                Label(self.pencere, text="Arıza Skor :", font="Verdana 12").place(x=175, y=325)
                Label(self.pencere, text=n.arızaSkor, font="Verdana 12").place(x=310, y=325)

        self.ABGAnasayfaButon = Button(self.pencere, text="Anasayfa", width=8,
        command=self.anasayfaGidiş, background="blue").place(x=480, y=400)
        self.ABGÇıkışButon = Button(self.pencere, text="Çıkış", width=5, background="red",
        command=self.pencere.quit).place(x=15, y=400)

    def usta_alan_bilgi_göster_kontrol(self):
        arızaGirdi = self.alanGirdi.get()
        sayac=0

        if arızaGirdi=="Elektrik" or arızaGirdi=="Mekanik" or arızaGirdi=="Hidrolik" or arızaGirdi=="Kalıp":
            sayac=1
            self.arızaGirdiDogru = tm.showinfo("Kontrol Mesajı", "Alan Ustaları Bulundu")
            self.usta_alan_bilgi_göster_sayfa(arızaGirdi)
        if sayac==0:
            tm.showerror("Kontrol Mesajı", "Alan Kaydı Bulunmadı")

    def usta_alan_bilgi_göster_sayfa(self, arızaGirdi):
        self.yokedici()
        self.resim1 = PhotoImage(file="Turkiye.png")
        Label(self.pencere, image=self.resim1).place(x=0, y=0)

        self.resim2 = PhotoImage(file="Fransa.png")
        Label(self.pencere, image=self.resim2).place(x=425, y=0)

        if arızaGirdi == "Mekanik":
            Label(text="Mekanik Ustaları", font="Verdana 15", fg="red").place(x=210, y=180)
        if arızaGirdi == "Elektrik":
            Label(text="Elektrik Ustaları", font="Verdana 15",fg="red").place(x=210, y=180)
        if arızaGirdi == "Hidrolik":
            Label(text="Hidrolik Ustaları", font="Verdana 15",fg="red").place(x=210, y=180)
        if arızaGirdi == "Kalıp":
            Label(text="Kalıp Ustaları", font="Verdana 15",fg="red").place(x=210, y=180)

        sayacY = 0

        for n in ustalarList:
            if n.alan == arızaGirdi:
                yKonum=220 + sayacY
                xKonum=50
                Label(text="ID", font="Verdana 10",).place(x=0+xKonum, y=yKonum)
                Label(text = n.id, font = "Verdana 10").place(x=20+xKonum , y=yKonum)
                Label(text="İsim", font="Verdana 10").place(x=40+xKonum, y=yKonum)
                Label(text = n.isim, font = "Verdana 10").place(x= 75+xKonum, y=yKonum)
                Label(text="Soyisim", font="Verdana 10").place(x=150+xKonum, y=yKonum)
                Label(text = n.soyisim, font ="Verdana 10").place(x=210+xKonum, y=yKonum)
                Label(text="Alan", font="Verdana 10").place(x=280+xKonum, y=yKonum)
                Label(text = n.alan, font = "Verdana 10").place(x=320+xKonum, y=yKonum)
                Label(text="Skor", font="Verdana 10").place(x=390+xKonum, y=yKonum)
                Label(text=n.skor, font="Verdana 10").place(x=430+xKonum, y=yKonum)
                sayacY = sayacY + 30

        self.UABGAnasayfaButon = Button(self.pencere, text="Anasayfa", width=8,
        command=self.anasayfaGidiş, background="blue").place(x=480, y=400)
        self.UABGÇıkışButon = Button(self.pencere, text="Çıkış", width=5, background="red",
        command=self.pencere.quit).place(x=15, y=400)

    def skorlama_kaydet(self):
        zaman = datetime.datetime.now()
        dosyaİsim = str(zaman.year) + str(zaman.month) + str(zaman.day) + str(zaman.hour) + str(zaman.minute) + ".xlsx"
        print(dosyaİsim)
        workbook = xlsxwriter.Workbook(dosyaİsim)
        worksheet = workbook.add_worksheet("Skorlar")

        row = 0
        col = 0

        for n in ustalarList:
            worksheet.write(row,col,n.id)
            worksheet.write(row,col+1,n.skor)
            row = row +1

        workbook.close()

    def skorlama_kontrol(self):
        idGirdi = self.skorİDGirdi.get()
        isimGirdi = self.skorİsimGirdi.get()
        soyisimGirdi = self.skorSoyisimGirdi.get()
        arızaGirdi = self.skorArızaGirdi.get()

        sayac = 0

        for n in ustalarList:
            if idGirdi == n.id:
                for m in arızalarList:
                    if m.arızaKodu == arızaGirdi:
                        if m.arızaAlan == n.alan:
                            sayac = 1
                            n.skor = m.arızaSkor + n.skor

        if sayac == 0:
            tm.showerror("Kontrol Mesajı", "Usta ve Arıza Alan Uyuşmazlığı")





    def anasayfa(self):
        self.resim1 = PhotoImage(file="Turkiye.png")
        Label(self.pencere, image=self.resim1).place(x=0, y=0)

        self.resim2 = PhotoImage(file="Fransa.png")
        Label(self.pencere, image=self.resim2).place(x=425, y=0)

        Label(text="Usta Bilgilerini Göster", font = "Verdana 10", fg="Blue").place(x=0,y=180)

        Label(text="İD", font="Verdana 10").place(x=0, y=220)
        self.idGirdi = Entry(self.pencere, width = 10)
        self.idGirdi.place(x=70, y=220)

        Label(text="İsim", font="Verdana 10").place(x=0, y=240)
        self.isimGirdi = Entry(self.pencere, width=10)
        self.isimGirdi.place(x=70, y=240)

        Label(text="Soyisim", font="Verdana 10").place(x=0, y=260)
        self.soyisimGirdi = Entry(self.pencere, width=10)
        self.soyisimGirdi.place(x=70, y=260)

        self.UBGButton = Button(self.pencere, text="Usta Bilgisi Göster", width=18,
        command =self.usta_bilgileri_sayfa_kontrol , background="blue").place(x=0, y=290)

        Label(text="Arıza Bilgileri Göster", font="Verdana 10", fg="red").place(x=200, y=180)

        Label(text="Arıza Kodu", font="Verdana 10").place(x=200, y=220)
        self.arızaKodGirdi = Entry(self.pencere,width=10)
        self.arızaKodGirdi.place(x=300, y=220)

        self.ABGButton = Button(self.pencere, text="Arıza Bilgisi Göster", width=18,
        command=self.arıza_bilgileri_sayfa_kontrol, background="blue").place(x=200,y=250)

        Label(text="Alan Ustaları Göster", font="Verdana 10", fg="green").place(x=400, y=180)

        Label(text="Alan", font="Verdana 10").place(x=400, y=220)
        self.alanGirdi = Entry(self.pencere, width=10)
        self.alanGirdi.place(x=470,y=220)

        self.UABGButton = Button(self.pencere, text="Alan Usta Göster", width=18,
        command =self.usta_alan_bilgi_göster_kontrol, background="blue").place(x=400,y=250)

        Label(text="Skorlama İşlemi", font ="Verdana 12", fg="purple").place(x=230,y=350)

        Label(text="İD", font="Verdana 10").place(x=230,y=380)
        self.skorİDGirdi = Entry(self.pencere, width=10)
        self.skorİDGirdi.place(x=300,y=380)

        Label(text="İsim", font="Verdana 10").place(x=230, y=410)
        self.skorİsimGirdi = Entry(self.pencere, width=10)
        self.skorİsimGirdi.place(x=300, y=410)

        Label(text="Soyisim", font="Verdana 10").place(x=230, y=440)
        self.skorSoyisimGirdi = Entry(self.pencere, width=10)
        self.skorSoyisimGirdi.place(x=300, y=440)

        Label(text="Arıza", font="Verdana 10").place(x=230, y=470)
        self.skorArızaGirdi = Entry(self.pencere, width=10)
        self.skorArızaGirdi.place(x=300, y=470)

        self.skorlaButton = Button(self.pencere, text="Skorla",width=18,
        command = self.skorlama_kontrol, background= "blue").place(x=230, y=500)

        self.skorKaydterButton = Button(self.pencere, text="Skorları Kaydet", width=18,
        command=self.skorlama_kaydet, background="blue").place(x=230, y=530)



uBM = excel_veri_çekme("Bilgiler.xlsx", "Ustalar")
aBM = excel_veri_çekme("Bilgiler.xlsx", "Arızalar")
kBM = excel_veri_çekme("Bilgiler.xlsx", "Kullanıcılar")

ustalarList, arızalarList, kullanıcılarList = [], [], []
for x in range(len(uBM)):
        yaş = zaman_hesaplama(uBM[x][3])
        işYıl = zaman_hesaplama(uBM[x][4])
        ustalarList.append(Usta(x+1,uBM[x][0],uBM[x][1],uBM[x][2],uBM[x][3],yaş,uBM[x][4],işYıl,uBM[x][5]))
for x in range(len(aBM)):
        arızalarList.append(Arıza(aBM[x][0],aBM[x][1],aBM[x][2],aBM[x][3]))
for x in range(len(kBM)):
        kullanıcılarList.append(Kullanıcı(kBM[x][0],kBM[x][1]))

print("Ustalar")
for n in ustalarList:
    print(n.id,n.isim,n.soyisim,n.alan,n.dTarih,n.yaş,n.bTarih,n.işYıl,n.skor,end="")
    print()
print("Arızalar")
for n in arızalarList:
    print(n.arızaKodu,n.arızaAlan,n.arızaSüre,n.arızaİsim,n.arızaSkor,end="")
    print()
print("Kullanıcılar")
for n in kullanıcılarList:
    print(n.kullanıcıİsim,n.şifre,end="")
    print()


pencere = Tk()
uygulama = Arayüz(pencere)
pencere.mainloop()